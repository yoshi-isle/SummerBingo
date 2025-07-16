#!/usr/bin/env python3
"""
Script to convert tile data from tiles.py into an Excel file that can be imported to Google Sheets.
"""

import sys
import os
import subprocess
from datetime import datetime

# Add the parent directory to sys.path to import from game_service_api
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'game_service_api'))

def install_dependencies():
    """Install required packages."""
    required_packages = ['openpyxl', 'xlsxwriter']
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print(f"📦 Installing missing packages: {', '.join(missing_packages)}")
        try:
            subprocess.check_call([
                sys.executable, '-m', 'pip', 'install'
            ] + missing_packages)
            print("✅ Dependencies installed successfully!")
            return True
        except subprocess.CalledProcessError as e:
            print(f"❌ Failed to install dependencies: {e}")
            return False
    
    return True

def create_excel_file():
    """Create an Excel file with all tile data."""
    if not install_dependencies():
        return False
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Import tile data
        from constants.tiles import world1_tiles, world2_tiles, world3_tiles, world4_tiles
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        world_header_font = Font(bold=True, size=14, color="FFFFFF")
        world_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_font = Font(bold=True, size=12, color="FFFFFF")
        section_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create overview sheet
        overview_sheet = wb.active
        overview_sheet.title = "Overview"
        
        # Add overview headers
        overview_headers = ["World", "Tile Type", "Tile ID", "Tile Name", "Description", "Required Count", "Wiki Link"]
        for col, header in enumerate(overview_headers, 1):
            cell = overview_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Data for all worlds
        worlds_data = [
            (world1_tiles, "World 1"),
            (world2_tiles, "World 2"),
            (world3_tiles, "World 3"),
            (world4_tiles, "World 4")
        ]
        
        current_row = 2
        
        # Process each world
        for world_data, world_name in worlds_data:
            # Create individual world sheet
            world_sheet = wb.create_sheet(title=world_name)
            world_row = 1
            
            # Add world title to world sheet
            world_sheet.merge_cells(f'A{world_row}:G{world_row}')
            world_title_cell = world_sheet.cell(row=world_row, column=1, value=world_name)
            world_title_cell.font = world_header_font
            world_title_cell.fill = world_header_fill
            world_title_cell.alignment = Alignment(horizontal='center', vertical='center')
            world_row += 1
            
            # Add headers to world sheet
            world_headers = ["Tile Type", "Tile ID", "Tile Name", "Description", "Required Count", "Wiki Link"]
            for col, header in enumerate(world_headers, 1):
                cell = world_sheet.cell(row=world_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            world_row += 1
            
            # Process each tile type
            tile_types = [
                ("world_tiles", "World Tiles"),
                ("key_tiles", "Key Tiles"),
                ("boss_tile", "Boss Tile")
            ]
            
            for tile_type, tile_type_name in tile_types:
                if tile_type not in world_data:
                    continue
                
                tiles = world_data[tile_type]
                if not tiles:
                    continue
                
                # Handle boss tile (single dict vs list)
                if tile_type == "boss_tile" and isinstance(tiles, dict):
                    tiles = [tiles]
                
                # Add section header to world sheet
                world_sheet.merge_cells(f'A{world_row}:F{world_row}')
                section_cell = world_sheet.cell(row=world_row, column=1, value=tile_type_name)
                section_cell.font = section_header_font
                section_cell.fill = section_header_fill
                section_cell.alignment = Alignment(horizontal='center', vertical='center')
                world_row += 1
                
                # Add tiles
                for tile in tiles:
                    tile_id = tile.get('id', 'N/A')
                    tile_name = tile.get('tile_name', 'Unknown').replace('\n', ' ')
                    description = tile.get('description', '').replace('\n', ' ')
                    completion_counter = tile.get('completion_counter', 1)
                    wiki_url = tile.get('wiki_url', '')
                    
                    # Add to overview sheet
                    overview_sheet.cell(row=current_row, column=1, value=world_name)
                    overview_sheet.cell(row=current_row, column=2, value=tile_type_name)
                    overview_sheet.cell(row=current_row, column=3, value=tile_id)
                    overview_sheet.cell(row=current_row, column=4, value=tile_name)
                    overview_sheet.cell(row=current_row, column=5, value=description)
                    overview_sheet.cell(row=current_row, column=6, value=completion_counter)
                    overview_sheet.cell(row=current_row, column=7, value=wiki_url)
                    
                    # Add borders to overview sheet
                    for col in range(1, 8):
                        overview_sheet.cell(row=current_row, column=col).border = thin_border
                    
                    # Add to world sheet
                    world_sheet.cell(row=world_row, column=1, value=tile_type_name)
                    world_sheet.cell(row=world_row, column=2, value=tile_id)
                    world_sheet.cell(row=world_row, column=3, value=tile_name)
                    world_sheet.cell(row=world_row, column=4, value=description)
                    world_sheet.cell(row=world_row, column=5, value=completion_counter)
                    world_sheet.cell(row=world_row, column=6, value=wiki_url)
                    
                    # Add borders to world sheet
                    for col in range(1, 7):
                        world_sheet.cell(row=world_row, column=col).border = thin_border
                    
                    current_row += 1
                    world_row += 1
                
                # Add spacing between sections
                world_row += 1
            
            # Auto-adjust column widths for world sheet
            for column in world_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                world_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Auto-adjust column widths for overview sheet
        for column in overview_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            overview_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        summary_sheet = wb.create_sheet(title="Summary")
        
        # Add summary title
        summary_sheet.merge_cells('A1:D1')
        summary_title = summary_sheet.cell(row=1, column=1, value="Summer Bingo Tile Summary")
        summary_title.font = world_header_font
        summary_title.fill = world_header_fill
        summary_title.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add generation date
        gen_date = summary_sheet.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
        gen_date.font = Font(italic=True)
        
        # Add summary headers
        summary_headers = ["World", "World Tiles", "Key Tiles", "Boss Tiles", "Total"]
        for col, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Calculate and add summary data
        total_world_tiles = 0
        total_key_tiles = 0
        total_boss_tiles = 0
        
        for i, (world_data, world_name) in enumerate(worlds_data, 1):
            world_tiles_count = len(world_data.get('world_tiles', []))
            key_tiles_count = len(world_data.get('key_tiles', []))
            boss_tiles_count = 1 if world_data.get('boss_tile') else 0
            world_total = world_tiles_count + key_tiles_count + boss_tiles_count
            
            total_world_tiles += world_tiles_count
            total_key_tiles += key_tiles_count
            total_boss_tiles += boss_tiles_count
            
            row = 4 + i
            summary_sheet.cell(row=row, column=1, value=world_name)
            summary_sheet.cell(row=row, column=2, value=world_tiles_count)
            summary_sheet.cell(row=row, column=3, value=key_tiles_count)
            summary_sheet.cell(row=row, column=4, value=boss_tiles_count)
            summary_sheet.cell(row=row, column=5, value=world_total)
            
            # Add borders
            for col in range(1, 6):
                summary_sheet.cell(row=row, column=col).border = thin_border
        
        # Add totals row
        totals_row = 4 + len(worlds_data) + 1
        summary_sheet.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
        summary_sheet.cell(row=totals_row, column=2, value=total_world_tiles).font = Font(bold=True)
        summary_sheet.cell(row=totals_row, column=3, value=total_key_tiles).font = Font(bold=True)
        summary_sheet.cell(row=totals_row, column=4, value=total_boss_tiles).font = Font(bold=True)
        summary_sheet.cell(row=totals_row, column=5, value=total_world_tiles + total_key_tiles + total_boss_tiles).font = Font(bold=True)
        
        # Add borders to totals row
        for col in range(1, 6):
            cell = summary_sheet.cell(row=totals_row, column=col)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Auto-adjust column widths for summary sheet
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        output_file = os.path.join(os.path.dirname(__file__), 'summer_bingo_tiles.xlsx')
        wb.save(output_file)
        
        print(f"✅ Excel file generated successfully: {output_file}")
        print(f"📊 Summary:")
        print(f"   - World tiles: {total_world_tiles}")
        print(f"   - Key tiles: {total_key_tiles}")
        print(f"   - Boss tiles: {total_boss_tiles}")
        print(f"   - Total tiles: {total_world_tiles + total_key_tiles + total_boss_tiles}")
        print(f"📄 Sheets created: Overview, Summary, {', '.join([world_name for _, world_name in worlds_data])}")
        print(f"📋 Ready to import to Google Sheets!")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function to create the Excel file."""
    try:
        if create_excel_file():
            return 0
        else:
            return 1
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        return 1

if __name__ == "__main__":
    exit(main())
